import logging
import inspect

from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import json
import csv


class Utils:

    def assretTextCompare(self, expected, actual):
        assert (expected.lower() == actual.lower())
        print("Assert pass")

    def custom_logger(logLevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        fh = logging.FileHandler("../automation.log", mode='w')
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s : %(message)s')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger
        pass

    def read_xlsx_test_data(self, file_name, sheet_name):
        test_data = []
        try:
            workbook = load_workbook(filename=file_name)
            sheet = workbook[sheet_name]

            headers = [sheet[get_column_letter(col)][0].value for col in range(1, sheet.max_column + 1)]

            for row in range(2, sheet.max_row + 1):
                data = {}
                for col, header in enumerate(headers, start=1):
                    data[header] = sheet[get_column_letter(col) + str(row)].value
                    print(data[header])
                test_data.append(data)
        except Exception as e:
            print(e)
            print("Error loading xlsx data")

        return test_data

    def read_json_test_data(self, path):
        test_data = []
        with open(path, "r") as file:
            test_data = json.load(file)
        return test_data

    def read_csv_test_data(self, path):
        test_data = []
        with open(path, 'r') as file:
            # csvreader = csv.reader(file, delimiter=',')
            csvreader = csv.reader(file)
            hearder = next(csvreader)  # gets the first line / row 1
            for row in csvreader:
                data = {}
                for i in range(0, len(hearder)):
                    data[hearder[i]] = row[i]
                    print(data[hearder[i]])
                test_data.append(data)
        return test_data


